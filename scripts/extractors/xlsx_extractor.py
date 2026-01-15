"""
XLSX Extractor for Course Materials

Extracts data from Excel files (formulas, templates, analytics tables).
"""

from pathlib import Path
from typing import Any


def extract_xlsx(file_path: Path) -> dict:
    """
    Extract content from an Excel file.

    Args:
        file_path: Path to the XLSX file

    Returns:
        Dictionary with sheets and their data
    """
    import pandas as pd
    from openpyxl import load_workbook

    result = {
        'sheets': {},
        'formulas': [],
        'summary': {}
    }

    # Load with openpyxl for formulas
    wb = load_workbook(file_path, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Extract formulas
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    result['formulas'].append({
                        'sheet': sheet_name,
                        'cell': cell.coordinate,
                        'formula': cell.value
                    })

    wb.close()

    # Load with pandas for data
    try:
        excel_file = pd.ExcelFile(file_path)

        for sheet_name in excel_file.sheet_names:
            try:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)

                # Clean data
                df = df.dropna(how='all', axis=0)
                df = df.dropna(how='all', axis=1)

                if df.empty:
                    continue

                # Convert to list of dicts
                # Clean column names
                df.columns = [str(col).strip() if pd.notna(col) else f'Column_{i}'
                             for i, col in enumerate(df.columns)]

                # Convert to records, handling NaN
                records = df.head(50).fillna('').to_dict('records')

                result['sheets'][sheet_name] = records

                # Add sheet summary
                result['summary'][sheet_name] = {
                    'rows': len(df),
                    'columns': list(df.columns),
                    'sample_values': {col: str(df[col].iloc[0]) if len(df) > 0 else ''
                                     for col in list(df.columns)[:5]}
                }

            except Exception as e:
                result['sheets'][sheet_name] = {'error': str(e)}

    except Exception as e:
        result['error'] = str(e)

    return result


def extract_formulas_summary(file_path: Path) -> list:
    """
    Extract only formulas from Excel file.
    Useful for understanding calculation logic.
    """
    from openpyxl import load_workbook

    formulas = []
    wb = load_workbook(file_path, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    # Try to get a meaningful name for the formula
                    row_header = ws.cell(row=cell.row, column=1).value
                    col_header = ws.cell(row=1, column=cell.column).value

                    formulas.append({
                        'sheet': sheet_name,
                        'cell': cell.coordinate,
                        'formula': cell.value,
                        'row_context': str(row_header) if row_header else '',
                        'col_context': str(col_header) if col_header else ''
                    })

    wb.close()
    return formulas
