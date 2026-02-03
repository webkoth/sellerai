#!/usr/bin/env python3
"""
Генератор отчёта НПД (налог на профессиональный доход) для Wildberries.

Скачивает документы из ЛК WB и рассчитывает налог:
- Еженедельные отчёты реализации (физ.лица) → 4%
- Уведомления о выкупе (юр.лица) → 6%

Использование:
    python scripts/generate_npd_report.py --year 2025
    python scripts/generate_npd_report.py --year 2025 --output reports/npd_2025.xlsx
"""

import argparse
import base64
import io
import json
import os
import re
import sys
import time
import zipfile
from datetime import datetime
from typing import Dict, List, Optional

try:
    import pdfplumber
except ImportError:
    print("Установите pdfplumber: pip install pdfplumber")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("Установите openpyxl: pip install openpyxl")
    sys.exit(1)

import urllib.request
import urllib.error

DOCUMENTS_API_URL = "https://documents-api.wildberries.ru"
RATE_LIMIT_DELAY = 10
NPD_LIMIT = 2_400_000


def get_token() -> str:
    token = os.environ.get('WB_API_TOKEN')
    if token:
        return token
    env_path = os.path.join(os.path.dirname(__file__), '..', '.env')
    if os.path.exists(env_path):
        with open(env_path) as f:
            for line in f:
                if line.startswith('WB_API_TOKEN='):
                    return line.split('=', 1)[1].strip().strip('"\'')
    raise ValueError("WB_API_TOKEN не найден")


def fetch_documents_list(token: str, date_from: str, date_to: str, limit: int = 50, offset: int = 0) -> dict:
    url = f"{DOCUMENTS_API_URL}/api/v1/documents/list?locale=ru&limit={limit}&offset={offset}&beginTime={date_from}&endTime={date_to}&sort=date&order=asc"
    req = urllib.request.Request(url, headers={"Authorization": token})
    with urllib.request.urlopen(req, timeout=60) as resp:
        return json.loads(resp.read())


def download_document(token: str, service_name: str, extension: str = "zip") -> bytes:
    url = f"{DOCUMENTS_API_URL}/api/v1/documents/download?serviceName={service_name}&extension={extension}"
    req = urllib.request.Request(url, headers={"Authorization": token})
    with urllib.request.urlopen(req, timeout=60) as resp:
        data = json.loads(resp.read())
    doc_b64 = data.get('data', {}).get('document', '')
    if not doc_b64:
        raise ValueError("Документ не содержит данных")
    return base64.b64decode(doc_b64)


def extract_from_zip(zip_bytes: bytes, extension: str) -> Optional[bytes]:
    try:
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
            for name in zf.namelist():
                if name.endswith(f'.{extension}'):
                    return zf.read(name)
    except:
        pass
    return None


def parse_weekly_report_pdf(pdf_bytes: bytes) -> float:
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            tables = pdf.pages[0].extract_tables()
            if not tables:
                return 0
            for table in tables:
                for row in table:
                    if not row or len(row) < 6:
                        continue
                    cell0 = str(row[0] or '').strip()
                    cell1 = str(row[1] or '').lower()
                    if (cell0 in ('1.', '1.1') and 'всего стоимость реализованного товара' in cell1):
                        amount_str = str(row[5] or '').replace(' ', '').replace(',', '.')
                        try:
                            return float(amount_str)
                        except:
                            pass
    except Exception as e:
        print(f"    Ошибка парсинга PDF: {e}", file=sys.stderr)
    return 0


def parse_redeem_notification_xlsx(xlsx_bytes: bytes) -> float:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            cell0 = str(row[0] or '').lower().strip()
            if cell0 in ('итого:', 'итого'):
                if len(row) > 4 and row[4]:
                    amount_str = str(row[4]).replace(' ', '').replace(',', '.')
                    try:
                        return float(amount_str)
                    except:
                        pass
    except Exception as e:
        print(f"    Ошибка парсинга XLSX: {e}", file=sys.stderr)
    return 0


def get_month_from_name(name: str, creation_time: str) -> str:
    match = re.search(r'(\d{4}-\d{2})-\d{2}', name)
    if match:
        return match.group(1)
    return creation_time[:7]


def generate_npd_report(year: int, verbose: bool = True) -> Dict:
    token = get_token()
    date_from = f"{year}-01-01"
    date_to = f"{year}-12-31"

    if verbose:
        print(f"Получение документов за {year} год...")

    all_documents = []
    offset = 0
    limit = 50

    while True:
        try:
            result = fetch_documents_list(token, date_from, date_to, limit, offset)
            docs = result.get('data', {}).get('documents', [])
            all_documents.extend(docs)
            if verbose:
                print(f"  Загружено {len(all_documents)} документов...")
            if len(docs) < limit:
                break
            offset += limit
            time.sleep(1)
        except urllib.error.HTTPError as e:
            if e.code == 429:
                if verbose:
                    print("  Rate limit, ждём 30 сек...")
                time.sleep(30)
            else:
                raise

    weekly_reports = [d for d in all_documents if 'еженедельный' in d.get('category', '').lower()]
    redeem_notifications = [d for d in all_documents if 'выкуп' in d.get('category', '').lower()]

    if verbose:
        print(f"\nНайдено: Еженедельные отчёты: {len(weekly_reports)}, Уведомления о выкупе: {len(redeem_notifications)}")

    months_data: Dict[str, Dict] = {}
    errors: List[str] = []

    if verbose:
        print(f"\n=== Еженедельные отчёты (НПД 4%) ===")

    for i, doc in enumerate(weekly_reports):
        name = doc['name'][:45]
        service_name = doc['serviceName']
        month = get_month_from_name(doc['name'], doc.get('creationTime', ''))

        if verbose:
            print(f"[{i+1:2}/{len(weekly_reports)}] {name}...", end=" ", flush=True)

        try:
            zip_bytes = download_document(token, service_name, 'zip')
            pdf_bytes = extract_from_zip(zip_bytes, 'pdf')

            if not pdf_bytes:
                if verbose:
                    print("нет PDF")
                errors.append(f"{name}: нет PDF")
                time.sleep(RATE_LIMIT_DELAY)
                continue

            amount = parse_weekly_report_pdf(pdf_bytes)

            if amount > 0:
                if month not in months_data:
                    months_data[month] = {'physical': 0, 'legal': 0}
                months_data[month]['physical'] += amount
                if verbose:
                    print(f"{amount:>12,.2f} ₽")
            else:
                if verbose:
                    print("не найдено")
                errors.append(f"{name}: не найдено")
        except Exception as e:
            if verbose:
                print(f"ошибка: {e}")
            errors.append(f"{name}: {e}")

        time.sleep(RATE_LIMIT_DELAY)

    if verbose:
        print(f"\n=== Уведомления о выкупе (НПД 6%) ===")

    for i, doc in enumerate(redeem_notifications):
        name = doc['name'][:45]
        service_name = doc['serviceName']
        month = get_month_from_name(doc['name'], doc.get('creationTime', ''))

        if verbose:
            print(f"[{i+1:2}/{len(redeem_notifications)}] {name}...", end=" ", flush=True)

        try:
            zip_bytes = download_document(token, service_name, 'zip')
            xlsx_bytes = extract_from_zip(zip_bytes, 'xlsx')

            if not xlsx_bytes:
                if verbose:
                    print("нет XLSX")
                errors.append(f"{name}: нет XLSX")
                time.sleep(RATE_LIMIT_DELAY)
                continue

            amount = parse_redeem_notification_xlsx(xlsx_bytes)

            if amount > 0:
                if month not in months_data:
                    months_data[month] = {'physical': 0, 'legal': 0}
                months_data[month]['legal'] += amount
                if verbose:
                    print(f"{amount:>12,.2f} ₽")
            else:
                if verbose:
                    print("не найдено")
                errors.append(f"{name}: не найдено")
        except Exception as e:
            if verbose:
                print(f"ошибка: {e}")
            errors.append(f"{name}: {e}")

        time.sleep(RATE_LIMIT_DELAY)

    total_physical = sum(m.get('physical', 0) for m in months_data.values())
    total_legal = sum(m.get('legal', 0) for m in months_data.values())
    total_physical_tax = total_physical * 0.04
    total_legal_tax = total_legal * 0.06
    total_tax = total_physical_tax + total_legal_tax
    total_income = total_physical + total_legal

    return {
        'year': year,
        'months': months_data,
        'totals': {
            'physical': total_physical,
            'legal': total_legal,
            'physical_tax': total_physical_tax,
            'legal_tax': total_legal_tax,
            'total_tax': total_tax,
            'total_income': total_income,
        },
        'limit': {
            'value': NPD_LIMIT,
            'remaining': NPD_LIMIT - total_income,
            'exceeded': total_income > NPD_LIMIT,
        },
        'errors': errors,
    }


def print_report(report: Dict):
    year = report['year']
    months = report['months']
    totals = report['totals']
    limit_info = report['limit']

    month_names = {
        '01': 'Январь', '02': 'Февраль', '03': 'Март', '04': 'Апрель',
        '05': 'Май', '06': 'Июнь', '07': 'Июль', '08': 'Август',
        '09': 'Сентябрь', '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'
    }

    print()
    print("=" * 80)
    print(f"                     ОТЧЁТ НПД ЗА {year} ГОД")
    print("=" * 80)
    print()
    print("| Месяц     | Физ.лица (4%)  | Налог 4%   | Юр.лица (6%) | Налог 6%  | Итого     |")
    print("|-----------|----------------|------------|--------------|-----------|-----------|")

    for month in sorted(months.keys()):
        if not month.startswith(str(year)):
            continue
        m = months[month]
        phys = m.get('physical', 0)
        legal = m.get('legal', 0)
        tax_phys = phys * 0.04
        tax_legal = legal * 0.06
        month_num = month.split('-')[1]
        month_name = month_names.get(month_num, month)
        print(f"| {month_name:9} | {phys:>14,.2f} | {tax_phys:>10,.2f} | {legal:>12,.2f} | {tax_legal:>9,.2f} | {tax_phys+tax_legal:>9,.2f} |")

    print("|-----------|----------------|------------|--------------|-----------|-----------|")
    print(f"| {'ИТОГО':9} | {totals['physical']:>14,.2f} | {totals['physical_tax']:>10,.2f} | {totals['legal']:>12,.2f} | {totals['legal_tax']:>9,.2f} | {totals['total_tax']:>9,.2f} |")

    print()
    print(f"📊 Физ.лица (4%): {totals['physical']:,.2f} ₽ → налог {totals['physical_tax']:,.2f} ₽")
    print(f"📊 Юр.лица (6%): {totals['legal']:,.2f} ₽ → налог {totals['legal_tax']:,.2f} ₽")
    print(f"💰 ОБЩИЙ НПД {year}: {totals['total_tax']:,.2f} ₽")

    print()
    if limit_info['exceeded']:
        print(f"⚠️ ВНИМАНИЕ: Доход {totals['total_income']:,.2f} ₽ ПРЕВЫСИЛ лимит НПД {limit_info['value']:,.2f} ₽!")
    else:
        remaining_pct = limit_info['remaining'] / limit_info['value'] * 100
        print(f"✅ До лимита НПД: {limit_info['remaining']:,.2f} ₽ ({remaining_pct:.1f}%)")


def save_to_excel(report: Dict, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"НПД {report['year']}"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    money_format = '#,##0.00'

    ws['A1'] = f"Отчёт НПД за {report['year']} год"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    headers = ['Месяц', 'Физ.лица (4%)', 'Налог 4%', 'Юр.лица (6%)', 'Налог 6%', 'Итого налог']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    month_names = {
        '01': 'Январь', '02': 'Февраль', '03': 'Март', '04': 'Апрель',
        '05': 'Май', '06': 'Июнь', '07': 'Июль', '08': 'Август',
        '09': 'Сентябрь', '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'
    }

    row = 4
    for month in sorted(report['months'].keys()):
        if not month.startswith(str(report['year'])):
            continue
        m = report['months'][month]
        phys = m.get('physical', 0)
        legal = m.get('legal', 0)
        tax_phys = phys * 0.04
        tax_legal = legal * 0.06
        month_num = month.split('-')[1]
        month_name = month_names.get(month_num, month)

        ws.cell(row=row, column=1, value=month_name)
        ws.cell(row=row, column=2, value=phys).number_format = money_format
        ws.cell(row=row, column=3, value=tax_phys).number_format = money_format
        ws.cell(row=row, column=4, value=legal).number_format = money_format
        ws.cell(row=row, column=5, value=tax_legal).number_format = money_format
        ws.cell(row=row, column=6, value=tax_phys + tax_legal).number_format = money_format
        row += 1

    totals = report['totals']
    ws.cell(row=row, column=1, value='ИТОГО').font = header_font
    ws.cell(row=row, column=2, value=totals['physical']).number_format = money_format
    ws.cell(row=row, column=3, value=totals['physical_tax']).number_format = money_format
    ws.cell(row=row, column=4, value=totals['legal']).number_format = money_format
    ws.cell(row=row, column=5, value=totals['legal_tax']).number_format = money_format
    ws.cell(row=row, column=6, value=totals['total_tax']).number_format = money_format

    for col in range(1, 7):
        ws.cell(row=row, column=col).font = header_font

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    row += 2
    ws.cell(row=row, column=1, value='Общий доход:')
    ws.cell(row=row, column=2, value=totals['total_income']).number_format = money_format

    row += 1
    ws.cell(row=row, column=1, value='Общий НПД:')
    ws.cell(row=row, column=2, value=totals['total_tax']).number_format = money_format
    ws.cell(row=row, column=2).font = Font(bold=True, color="008000")

    row += 1
    limit_info = report['limit']
    ws.cell(row=row, column=1, value='Лимит НПД:')
    ws.cell(row=row, column=2, value=limit_info['value']).number_format = money_format

    row += 1
    ws.cell(row=row, column=1, value='Осталось до лимита:')
    cell = ws.cell(row=row, column=2, value=limit_info['remaining'])
    cell.number_format = money_format
    if limit_info['exceeded']:
        cell.font = Font(bold=True, color="FF0000")

    wb.save(output_path)
    print(f"\n📁 Отчёт сохранён: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Генератор отчёта НПД для Wildberries')
    parser.add_argument('--year', type=int, default=datetime.now().year, help='Год для расчёта')
    parser.add_argument('--output', '-o', type=str, help='Путь для сохранения Excel файла')
    parser.add_argument('--quiet', '-q', action='store_true', help='Минимальный вывод')

    args = parser.parse_args()

    report = generate_npd_report(args.year, verbose=not args.quiet)
    print_report(report)

    if args.output:
        save_to_excel(report, args.output)


if __name__ == '__main__':
    main()
