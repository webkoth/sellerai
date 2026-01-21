#!/usr/bin/env python3
"""
Генератор отчёта о прибыльности товаров Wildberries.

Использование:
    .venv/bin/python3 scripts/generate_profitability_report.py --cost-file reports/sebestoimost.csv --data-file reports/product_profitability.json

Или для генерации шаблона:
    .venv/bin/python3 scripts/generate_profitability_report.py --generate-template --output reports/cost_template.csv
"""

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Библиотека openpyxl не установлена")
    print("   Установите: .venv/bin/pip install openpyxl")
    sys.exit(1)


# Стили Excel
STYLES = {
    'title_font': Font(bold=True, size=16),
    'header_font': Font(bold=True, size=12),
    'header_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
    'header_font_white': Font(bold=True, color="FFFFFF"),
    'green_fill': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'red_fill': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'red_header': PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
    'thin_border': Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
}


def check_required_files(cost_file: str, data_file: str = None) -> dict:
    """Проверяет наличие необходимых файлов."""
    result = {
        'cost_file': None,
        'data_file': None,
        'errors': []
    }

    # Проверка файла себестоимости
    if cost_file:
        cost_path = Path(cost_file)
        if cost_path.exists():
            result['cost_file'] = cost_path
        else:
            result['errors'].append(f"Файл себестоимости не найден: {cost_file}")
    else:
        # Поиск файла по паттернам
        patterns = ['*sebestoimost*.csv', '*cost*.csv', '*себестоимость*.csv']
        reports_dir = Path('reports')
        for pattern in patterns:
            files = list(reports_dir.glob(pattern))
            if files:
                result['cost_file'] = files[0]
                break

        if not result['cost_file']:
            result['errors'].append("Файл себестоимости не найден. Укажите путь через --cost-file")

    # Проверка файла данных
    if data_file:
        data_path = Path(data_file)
        if data_path.exists():
            result['data_file'] = data_path
        else:
            result['errors'].append(f"Файл данных не найден: {data_file}")

    return result


def load_cost_data(cost_file: Path) -> dict:
    """Загружает данные о себестоимости из CSV."""
    costs = {}

    with open(cost_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Ищем колонку с артикулом
            article = row.get('Артикул продавца') or row.get('article') or row.get('sku')
            # Ищем колонку с себестоимостью
            cost = row.get('СЕБЕСТОИМОСТЬ') or row.get('себестоимость') or row.get('cost') or row.get('Себестоимость')

            if article and cost:
                try:
                    costs[article] = float(str(cost).replace(',', '.').replace(' ', ''))
                except ValueError:
                    pass

    return costs


def create_excel_report(data: dict, output_path: str):
    """Создаёт Excel отчёт."""
    summary = data['summary']
    products = data['products']
    products_sorted = sorted(products, key=lambda x: x['grossProfit'], reverse=True)

    wb = openpyxl.Workbook()

    # === ЛИСТ 1: СВОДКА ===
    ws_summary = wb.active
    ws_summary.title = "Сводка"

    ws_summary['A1'] = "ОТЧЁТ О ПРИБЫЛЬНОСТИ ТОВАРОВ"
    ws_summary['A1'].font = STYLES['title_font']
    ws_summary.merge_cells('A1:C1')
    ws_summary['A2'] = f"Дата формирования: {datetime.now().strftime('%d.%m.%Y')}"

    summary_rows = [
        ["Показатель", "Значение"],
        ["Товаров в отчёте", summary['totalProducts']],
        ["Продано единиц", summary['totalSold']],
        ["Выручка к перечислению", f"{summary['totalRevenue']:,.0f} ₽"],
        ["Общая себестоимость", f"{summary['totalCost']:,.0f} ₽"],
        ["Валовая прибыль", f"{summary['totalProfit']:,.0f} ₽"],
        ["Средняя маржа", f"{summary['avgMargin']:.1f}%"],
    ]

    for i, row in enumerate(summary_rows, start=4):
        for j, val in enumerate(row, start=1):
            cell = ws_summary.cell(row=i, column=j, value=val)
            if i == 4:
                cell.font = STYLES['header_font_white']
                cell.fill = STYLES['header_fill']

    # Структура портфеля
    profitable = [p for p in products if p['grossProfit'] >= 0]
    loss = [p for p in products if p['grossProfit'] < 0]
    profitable_sum = sum(p['grossProfit'] for p in profitable)
    loss_sum = sum(p['grossProfit'] for p in loss)

    ws_summary['A12'] = "Структура портфеля"
    ws_summary['A12'].font = STYLES['header_font']

    portfolio_rows = [
        ["Категория", "Кол-во", "Прибыль/Убыток"],
        ["🟢 Прибыльные", len(profitable), f"+{profitable_sum:,.0f} ₽"],
        ["🔴 Убыточные", len(loss), f"{loss_sum:,.0f} ₽"],
    ]

    for i, row in enumerate(portfolio_rows, start=13):
        for j, val in enumerate(row, start=1):
            cell = ws_summary.cell(row=i, column=j, value=val)
            if i == 13:
                cell.font = STYLES['header_font_white']
                cell.fill = STYLES['header_fill']
            elif i == 14:
                cell.fill = STYLES['green_fill']
            elif i == 15:
                cell.fill = STYLES['red_fill']

    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 20

    # === ЛИСТ 2: ВСЕ ТОВАРЫ ===
    ws_products = wb.create_sheet("Все товары")

    headers = ["#", "Артикул", "Название", "Категория", "Продано", "Розн. цена",
               "Комиссия WB", "К перечислению", "Себест./ед", "Себест. всего",
               "Прибыль", "Маржа", "ROI"]

    for j, header in enumerate(headers, start=1):
        cell = ws_products.cell(row=1, column=j, value=header)
        cell.font = STYLES['header_font_white']
        cell.fill = STYLES['header_fill']
        cell.border = STYLES['thin_border']
        cell.alignment = Alignment(horizontal='center')

    for i, p in enumerate(products_sorted, start=2):
        row_data = [
            i - 1, p['article'], p['name'], p.get('category', ''),
            p['sold'], p['retailEstimated'], p['commission'], p['toPay'],
            p['costPerUnit'], p['costTotal'], p['grossProfit'],
            p['margin'] / 100, p['roi'] / 100
        ]

        for j, val in enumerate(row_data, start=1):
            cell = ws_products.cell(row=i, column=j, value=val)
            cell.border = STYLES['thin_border']

            if j in [6, 7, 8, 9, 10, 11]:
                cell.number_format = '#,##0 ₽'
            elif j in [12, 13]:
                cell.number_format = '0.0%'

            if j == 11:
                cell.fill = STYLES['green_fill'] if val >= 0 else STYLES['red_fill']

    col_widths = [5, 22, 45, 25, 10, 14, 14, 16, 12, 14, 14, 10, 10]
    for i, width in enumerate(col_widths, start=1):
        ws_products.column_dimensions[get_column_letter(i)].width = width
    ws_products.freeze_panes = 'A2'

    # === ЛИСТ 3: ТОП-20 ===
    ws_top = wb.create_sheet("ТОП-20")

    for j, header in enumerate(headers, start=1):
        cell = ws_top.cell(row=1, column=j, value=header)
        cell.font = STYLES['header_font_white']
        cell.fill = STYLES['header_fill']
        cell.border = STYLES['thin_border']

    for i, p in enumerate(products_sorted[:20], start=2):
        row_data = [
            i - 1, p['article'], p['name'], p.get('category', ''),
            p['sold'], p['retailEstimated'], p['commission'], p['toPay'],
            p['costPerUnit'], p['costTotal'], p['grossProfit'],
            p['margin'] / 100, p['roi'] / 100
        ]
        for j, val in enumerate(row_data, start=1):
            cell = ws_top.cell(row=i, column=j, value=val)
            cell.border = STYLES['thin_border']
            if j in [6, 7, 8, 9, 10, 11]:
                cell.number_format = '#,##0 ₽'
            elif j in [12, 13]:
                cell.number_format = '0.0%'
            if j == 11:
                cell.fill = STYLES['green_fill']

    for i, width in enumerate(col_widths, start=1):
        ws_top.column_dimensions[get_column_letter(i)].width = width
    ws_top.freeze_panes = 'A2'

    # === ЛИСТ 4: УБЫТОЧНЫЕ ===
    ws_loss = wb.create_sheet("Убыточные")

    loss_headers = ["#", "Артикул", "Название", "Продано", "Розн. цена",
                    "Комиссия", "Себестоимость", "Убыток", "Маржа", "Рекомендация"]

    for j, header in enumerate(loss_headers, start=1):
        cell = ws_loss.cell(row=1, column=j, value=header)
        cell.font = STYLES['header_font_white']
        cell.fill = STYLES['red_header']
        cell.border = STYLES['thin_border']

    loss_products = [p for p in products_sorted if p['grossProfit'] < 0]
    for i, p in enumerate(sorted(loss_products, key=lambda x: x['grossProfit']), start=2):
        rec = "Поднять цену" if p['margin'] < -10 else "Пересмотреть"
        row_data = [
            i - 1, p['article'], p['name'], p['sold'],
            p['retailEstimated'], p['commission'], p['costTotal'],
            p['grossProfit'], p['margin'] / 100, rec
        ]
        for j, val in enumerate(row_data, start=1):
            cell = ws_loss.cell(row=i, column=j, value=val)
            cell.border = STYLES['thin_border']
            cell.fill = STYLES['red_fill']
            if j in [5, 6, 7, 8]:
                cell.number_format = '#,##0 ₽'
            elif j == 9:
                cell.number_format = '0.0%'

    loss_widths = [5, 22, 40, 10, 14, 12, 14, 12, 10, 16]
    for i, width in enumerate(loss_widths, start=1):
        ws_loss.column_dimensions[get_column_letter(i)].width = width

    # Сохранение
    wb.save(output_path)
    return len(products), len(loss_products)


def generate_cost_template(products: list, output_path: str):
    """Генерирует шаблон для заполнения себестоимости."""
    with open(output_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Артикул продавца', 'Название товара', 'Категория', 'nmId WB',
                        'Продано шт', 'К перечислению ₽', 'СЕБЕСТОИМОСТЬ'])

        for p in products:
            writer.writerow([
                p.get('article', ''),
                p.get('name', ''),
                p.get('category', ''),
                p.get('nmId', ''),
                p.get('sold', 0),
                p.get('toPay', 0),
                ''  # Пустая колонка для заполнения
            ])

    return len(products)


def main():
    parser = argparse.ArgumentParser(description='Генератор отчёта о прибыльности товаров WB')
    parser.add_argument('--cost-file', help='Путь к CSV файлу с себестоимостью')
    parser.add_argument('--data-file', help='Путь к JSON файлу с данными о продажах')
    parser.add_argument('--output', default='reports/product_profitability_report.xlsx',
                        help='Путь для сохранения отчёта')
    parser.add_argument('--generate-template', action='store_true',
                        help='Сгенерировать шаблон для заполнения себестоимости')
    parser.add_argument('--check', action='store_true',
                        help='Только проверить наличие необходимых файлов')

    args = parser.parse_args()

    # Режим проверки
    if args.check:
        result = check_required_files(args.cost_file, args.data_file)
        if result['errors']:
            print("❌ Не хватает данных:")
            for err in result['errors']:
                print(f"   - {err}")
            sys.exit(1)
        else:
            print("✅ Все необходимые файлы найдены:")
            print(f"   - Себестоимость: {result['cost_file']}")
            if result['data_file']:
                print(f"   - Данные: {result['data_file']}")
            sys.exit(0)

    # Режим генерации шаблона
    if args.generate_template:
        if not args.data_file:
            print("❌ Для генерации шаблона укажите --data-file с данными о товарах")
            sys.exit(1)

        with open(args.data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)

        output = args.output if args.output.endswith('.csv') else 'reports/cost_template.csv'
        count = generate_cost_template(data['products'], output)
        print(f"✅ Шаблон создан: {output}")
        print(f"   Товаров: {count}")
        print(f"\n💡 Заполните колонку СЕБЕСТОИМОСТЬ и запустите отчёт снова")
        sys.exit(0)

    # Режим генерации отчёта
    if not args.data_file:
        print("❌ Укажите путь к JSON файлу с данными: --data-file")
        sys.exit(1)

    with open(args.data_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    total, loss = create_excel_report(data, args.output)

    summary = data['summary']
    print(f"✅ Отчёт создан: {args.output}")
    print(f"\n📊 Ключевые показатели:")
    print(f"   Товаров: {summary['totalProducts']}")
    print(f"   Валовая прибыль: {summary['totalProfit']:,.0f}₽")
    print(f"   Маржа: {summary['avgMargin']:.1f}%")
    print(f"   🟢 Прибыльных: {summary['profitableCount']}")
    print(f"   🔴 Убыточных: {loss}")


if __name__ == '__main__':
    main()
